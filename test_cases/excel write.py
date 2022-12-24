import openpyxl

path="C:\projects\pythonProject\Co_Repo.xlsx"
wb=openpyxl.load_workbook(path)
sheet=wb["Sheet1"]

sn=wb.sheetnames

print(sn)
# sheet=wb.active
rows=sheet.max_row
cols=sheet.max_column
print(rows)
print(cols)


for r in range(1, rows+1):
    for c in range(1, cols+1):
        print(sheet.cell(r,c).value, end="  ")
    print()
